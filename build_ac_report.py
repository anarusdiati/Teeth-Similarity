"""
Bangun laporan penilaian AC: XLSX dengan thumbnail tertanam + CSV pendamping.

Input : judgments.json  — daftar {"file", "grade", "note"} (urutan = nomor urut)
        folder gambar
Output: ac_report.xlsx  — kolom: Nomor | Nama File | Gambar | Grade AC | Note
        ac_report.csv   — sama, tapi kolom Gambar berisi nama file (CSV tidak bisa
                          menyimpan gambar)

Pakai:
    python3 build_ac_report.py --img-dir "Front Teeth drg Laura" \
                               --judgments judgments.json \
                               --out ac_report
"""

from __future__ import annotations

import argparse
import csv
import json
import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

try:                                    # .heic butuh plugin; opsional
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HEIC_OK = True
except Exception:
    HEIC_OK = False

THUMB_W = 260                           # lebar thumbnail (px)
ROW_H_PT = 150                          # tinggi baris (pt)
FONT = "Arial"

# Banding rujukan AC (Evans & Shaw): 1-4 tidak butuh, 5-7 borderline, 8-10 butuh.
BAND_FILL = {
    "tidak butuh perawatan": PatternFill("solid", fgColor="E8F5E9"),
    "borderline":            PatternFill("solid", fgColor="FFF8E1"),
    "butuh perawatan":       PatternFill("solid", fgColor="FFEBEE"),
}


def band_of(grade: int) -> str:
    if grade <= 4:
        return "tidak butuh perawatan"
    if grade <= 7:
        return "borderline"
    return "butuh perawatan"


def make_thumb(path: str, width: int = THUMB_W) -> tuple[BytesIO, int, int]:
    """Resize ke PNG in-memory. Selalu konversi supaya .heic/.JPG seragam."""
    im = Image.open(path).convert("RGB")
    ratio = width / im.width
    im = im.resize((width, max(1, int(im.height * ratio))), Image.LANCZOS)
    buf = BytesIO()
    im.save(buf, format="PNG")
    buf.seek(0)
    return buf, im.width, im.height


def build(img_dir: str, judgments: list[dict], out_base: str) -> tuple[str, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Penilaian AC"

    headers = ["Nomor", "Nama File", "Gambar", "Grade AC", "Note"]
    widths = [8, 42, 38, 10, 95]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    csv_rows = []
    missing = []

    for i, j in enumerate(judgments, start=1):
        r = i + 1
        fname, grade, note = j["file"], int(j["grade"]), j["note"]
        band = band_of(grade)

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=fname)
        ws.cell(row=r, column=4, value=grade)
        ws.cell(row=r, column=5, value=note)

        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c in (1, 4) else "left",
                vertical="center", wrap_text=(c == 5))
        gc = ws.cell(row=r, column=4)
        gc.font = Font(name=FONT, size=12, bold=True)
        gc.fill = BAND_FILL[band]

        path = os.path.join(img_dir, fname)
        if os.path.exists(path):
            try:
                buf, w_px, h_px = make_thumb(path)
                xl = XLImage(buf)
                xl.width, xl.height = w_px, h_px
                ws.add_image(xl, f"C{r}")
                ws.row_dimensions[r].height = max(ROW_H_PT, h_px * 0.78)
            except Exception as e:                      # .heic tanpa plugin, file rusak
                ws.cell(row=r, column=3, value=f"[gagal dimuat: {e}]")
                ws.row_dimensions[r].height = ROW_H_PT
                missing.append(fname)
        else:
            ws.cell(row=r, column=3, value="[file tidak ditemukan]")
            ws.row_dimensions[r].height = ROW_H_PT
            missing.append(fname)

        csv_rows.append({"Nomor": i, "Nama File": fname, "Gambar": fname,
                         "Grade AC": grade, "Band": band, "Note": note})

    xlsx_path = f"{out_base}.xlsx"
    csv_path = f"{out_base}.csv"
    wb.save(xlsx_path)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["Nomor", "Nama File", "Gambar",
                                          "Grade AC", "Band", "Note"])
        w.writeheader()
        w.writerows(csv_rows)

    if missing:
        print(f"PERINGATAN: {len(missing)} gambar tidak termuat: {missing[:5]}")
        if not HEIC_OK and any(m.lower().endswith(".heic") for m in missing):
            print("  -> ada .heic tapi pillow_heif belum terpasang: pip install pillow-heif")
    return xlsx_path, csv_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--img-dir", required=True)
    ap.add_argument("--judgments", required=True)
    ap.add_argument("--out", default="ac_report")
    a = ap.parse_args()

    with open(a.judgments, encoding="utf-8") as f:
        data = json.load(f)
    x, c = build(a.img_dir, data, a.out)
    print(f"{len(data)} baris -> {x} , {c}")
