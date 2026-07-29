"""
Bangun satu workbook AC dengan sheet Train / Val / Test + CSV per split.

Grade 0 = DIKECUALIKAN (foto tidak valid: bukan tampak frontal, dsb).
Baris grade 0 tetap ditulis agar terlacak, tapi ditandai dan tidak masuk statistik.

Pakai:
    python3 build_ac_workbook.py
"""

from __future__ import annotations

import csv
import json
import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

BASE_IMG = "/sessions/relaxed-beautiful-cray/mnt/OMNI Front Teeth"
OUT = "/sessions/relaxed-beautiful-cray/mnt/Teeth-Similarity"

SPLITS = [
    ("Train", "Train-Front-Without Braces", "judgments_train_claude.json"),
    ("Val",   "Val-Front-Without Braces",   "judgments_val_claude.json"),
    ("Test",  "Test-Front-Without Braces",  "judgments_test_claude.json"),
]

THUMB_W = 260
ROW_H_PT = 150
FONT = "Arial"

BAND_FILL = {
    "tidak butuh perawatan": PatternFill("solid", fgColor="E8F5E9"),
    "borderline":            PatternFill("solid", fgColor="FFF8E1"),
    "butuh perawatan":       PatternFill("solid", fgColor="FFEBEE"),
    "DIKECUALIKAN":          PatternFill("solid", fgColor="E0E0E0"),
}


def band_of(grade: int) -> str:
    if grade == 0:
        return "DIKECUALIKAN"
    if grade <= 4:
        return "tidak butuh perawatan"
    if grade <= 7:
        return "borderline"
    return "butuh perawatan"


def make_thumb(path: str, width: int = THUMB_W):
    im = Image.open(path).convert("RGB")
    im = im.resize((width, max(1, int(im.height * width / im.width))), Image.LANCZOS)
    buf = BytesIO()
    im.save(buf, format="PNG")
    buf.seek(0)
    return buf, im.width, im.height


def fill_sheet(ws, img_dir: str, judgments: list[dict]) -> list[dict]:
    headers = ["Nomor", "Nama File", "Gambar", "Grade AC", "Note"]
    widths = [8, 30, 38, 10, 100]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    rows = []
    for i, j in enumerate(judgments, start=1):
        r = i + 1
        fname, grade, note = j["file"], int(j["grade"]), j["note"]
        band = band_of(grade)

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=fname)
        ws.cell(row=r, column=4, value=(grade if grade else "—"))
        ws.cell(row=r, column=5, value=note)

        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c in (1, 4) else "left",
                vertical="center", wrap_text=(c == 5))
        gc = ws.cell(row=r, column=4)
        gc.font = Font(name=FONT, size=12, bold=True)
        gc.fill = BAND_FILL[band]

        path = os.path.join(img_dir, fname)
        if os.path.exists(path):
            buf, w_px, h_px = make_thumb(path)
            xl = XLImage(buf)
            xl.width, xl.height = w_px, h_px
            ws.add_image(xl, f"C{r}")
            ws.row_dimensions[r].height = max(ROW_H_PT, h_px * 0.78)
        else:
            ws.cell(row=r, column=3, value="[file tidak ditemukan]")
            ws.row_dimensions[r].height = ROW_H_PT

        rows.append({"Nomor": i, "Nama File": fname, "Gambar": fname,
                     "Grade AC": grade, "Band": band, "Note": note})
    return rows


def main():
    wb = Workbook()
    wb.remove(wb.active)
    summary = {}

    for sheet_name, sub, jfile in SPLITS:
        jpath = os.path.join(OUT, jfile)
        ws = wb.create_sheet(sheet_name)

        if not os.path.exists(jpath):
            ws["A1"] = f"BELUM DINILAI — {sub}"
            ws["A1"].font = Font(name=FONT, bold=True, size=12)
            ws["A2"] = (f"Jalankan ulang setelah {jfile} tersedia. "
                        f"Struktur kolom sama dengan sheet lain.")
            ws["A2"].font = Font(name=FONT, size=10)
            summary[sheet_name] = None
            print(f"{sheet_name:6s} dilewati — {jfile} belum ada")
            continue

        with open(jpath, encoding="utf-8") as f:
            judgments = json.load(f)
        rows = fill_sheet(ws, os.path.join(BASE_IMG, sub), judgments)

        csv_path = os.path.join(OUT, f"ac_{sheet_name.lower()}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["Nomor", "Nama File", "Gambar",
                                              "Grade AC", "Band", "Note"])
            w.writeheader()
            w.writerows(rows)

        graded = [r["Grade AC"] for r in rows if r["Grade AC"] > 0]
        summary[sheet_name] = (len(rows), len(rows) - len(graded), graded)
        print(f"{sheet_name:6s} {len(rows):3d} baris "
              f"({len(rows)-len(graded)} dikecualikan) -> {csv_path}")

    out_xlsx = os.path.join(OUT, "ac_report.xlsx")
    wb.save(out_xlsx)
    print(f"\n{out_xlsx}")
    return summary


if __name__ == "__main__":
    main()
